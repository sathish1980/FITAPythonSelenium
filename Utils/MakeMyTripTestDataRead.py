import openpyxl


class MakeMyTripTestDataRead():

    @staticmethod
    def excelFileRead():
        Data = {}
        WBK = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\FITAPythonSelenium\\input\\MakeMyTrip.xlsx")
        sheetName = WBK["ValidSearch"]
        totalRows = sheetName.max_row
        print("totalrows: ",totalRows)
        totalColumn = sheetName.max_column
        print("totalColumn: ",totalColumn)
        for eachRow in range(1,totalRows+1):
            for eachColumn in range(1, totalColumn + 1):
                eachValue = sheetName.cell(row=eachRow,column=eachColumn).value
                if(eachColumn==1):
                    Data[sheetName.cell(row=1,column=eachColumn).value+str(eachRow)]=eachValue
                if (eachColumn == 2):
                    Data[sheetName.cell(row=1,column=eachColumn).value+str(eachRow)]=eachValue
                if (eachColumn == 3):
                    Data[sheetName.cell(row=1,column=eachColumn).value+str(eachRow)]=eachValue
        print(Data)
        WBK.close()
        return Data # Sending the dictonary in to list since in the conftest.py  we need to get a value in the list
