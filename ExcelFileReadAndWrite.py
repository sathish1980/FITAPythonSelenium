import openpyxl


class excelFile():
    Data={}
    def excelFileRead(self,Sheetname):
        WBK = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\FITAPythonSelenium\\input\\Login.xlsx")
        sheetName = WBK[Sheetname]
        totalRows = sheetName.max_row
        print("totalrows: ",totalRows)
        totalColumn = sheetName.max_column
        print("totalColumn: ",totalColumn)
        for eachRow in range(1,totalRows+1):
            for eachColumn in range(1, totalColumn + 1):
                eachValue = sheetName.cell(row=eachRow,column=eachColumn).value
                if(eachColumn==1):
                    self.Data[sheetName.cell(row=1,column=eachColumn).value+str(eachRow)]=eachValue
                if (eachColumn == 2):
                    self.Data[sheetName.cell(row=1,column=eachColumn).value+str(eachRow)]=eachValue
                if (eachColumn == 3):
                    self.Data[sheetName.cell(row=1,column=eachColumn).value+str(eachRow)]=eachValue
        print(self.Data)
        WBK.close()

    def writeInToExcel(self):
        outworkbook = openpyxl.Workbook()
        outputSheet = outworkbook.active
        outputSheet.cell(row=1,column=1).value="sathish"
        outputSheet.cell(row=1, column=2).value = "kumar"
        outputSheet.cell(row=1, column=3).value = "R"
        outworkbook.save("C:\\Users\\sathishkumar\\PycharmProjects\\FITAPythonSelenium\\Output\\output.xlsx")
        print("Write is done")
        outworkbook.close()

    def excelFileReadandWrite(self,Sheetname):
        WBK = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\FITAPythonSelenium\\input\\Login.xlsx")
        sheetName = WBK[Sheetname]
        totalRows = sheetName.max_row
        print("totalrows: ",totalRows)
        totalColumn = sheetName.max_column
        print("totalColumn: ",totalColumn)
        outworkbook = openpyxl.Workbook()
        outputSheet = outworkbook.active
        for eachRow in range(1,totalRows+1):
            for eachColumn in range(1, totalColumn + 1):
                eachValue = sheetName.cell(row=eachRow,column=eachColumn).value
                outputSheet.cell(row=eachRow, column=eachColumn).value=eachValue

        outworkbook.save("C:\\Users\\sathishkumar\\PycharmProjects\\FITAPythonSelenium\\Output\\output.xlsx")
        print("Write is done")
        outworkbook.close()
        WBK.close()

obj=excelFile()
#obj.excelFileRead("Credentials")
obj.excelFileReadandWrite("Credentials")